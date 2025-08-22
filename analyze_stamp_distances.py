#!/usr/bin/env python3
"""
Stamp Distance Analysis Tool

This script analyzes the distance between geocoded pilgrim stamps and the Camino Frances trail.
It identifies stamps that are positioned beyond a specified distance threshold from the trail.
"""

import pandas as pd
import json
import folium
from shapely.geometry import Point, LineString
from shapely.ops import nearest_points
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import warnings
import sys
import os
import math

warnings.filterwarnings('ignore')

def parse_arguments():
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(
        description='Analyze distances between pilgrim stamps and Camino Frances trail',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                    # Use default 5.0 km threshold
  %(prog)s -d 2.0            # Use 2.0 km threshold
  %(prog)s --distance 10.0   # Use 10.0 km threshold
  %(prog)s -f input.csv      # Use custom input file
        """
    )
    parser.add_argument(
        '-d', '--distance',
        type=float,
        default=5.0,
        help='Maximum allowed distance from trail in kilometers (default: 5.0)'
    )
    parser.add_argument(
        '-f', '--input-file',
        type=str,
        default='data/pilgrim_stamps_geocoded.csv',
        help='Input CSV file path (default: data/pilgrim_stamps_geocoded.csv)'
    )
    
    args = parser.parse_args()
    
    if args.distance <= 0:
        parser.error("Distance must be a positive number")
    
    return args

def load_trail_data(geojson_path):
    """Load and validate trail data from GeoJSON file"""
    try:
        with open(geojson_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if data['type'] != 'FeatureCollection':
            raise ValueError("Invalid GeoJSON: must be FeatureCollection")
        
        features = data['features']
        if not features:
            raise ValueError("No features found in GeoJSON")
        
        # Extract coordinates from the first feature (assuming single trail)
        coords = features[0]['geometry']['coordinates']
        if not coords or len(coords) < 2:
            raise ValueError("Invalid trail: insufficient coordinates")
        
        print(f"✓ Loaded trail with {len(coords)} coordinate points")
        return coords
        
    except FileNotFoundError:
        print(f"❌ Error: Trail file '{geojson_path}' not found")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"❌ Error: Invalid JSON in '{geojson_path}'")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error loading trail data: {e}")
        sys.exit(1)

def load_stamps_data(csv_path):
    """Load and validate stamps data from CSV file"""
    try:
        df = pd.read_csv(csv_path)
        
        required_columns = ['place', 'town', 'latitude', 'longitude']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            print(f"❌ Error: Missing required columns: {missing_columns}")
            sys.exit(1)
        
        # Check for valid coordinates
        invalid_coords = df[
            (df['latitude'].isna()) | 
            (df['longitude'].isna()) |
            (df['latitude'] == 0) | 
            (df['longitude'] == 0)
        ]
        
        if len(invalid_coords) > 0:
            print(f"⚠️  Warning: {len(invalid_coords)} stamps have invalid coordinates")
        
        print(f"✓ Loaded {len(df)} stamps from CSV")
        return df
        
    except FileNotFoundError:
        print(f"❌ Error: Stamps file '{csv_path}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error loading stamps data: {e}")
        sys.exit(1)

def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate the great circle distance between two points on Earth"""
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    
    # Haversine formula
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    
    # Radius of Earth in kilometers
    r = 6371
    
    return c * r

def calculate_distances_to_trail(stamps_df, trail_coords, max_distance_km):
    """Calculate distances from stamps to trail and identify those beyond threshold"""
    print(f"Calculating distances to trail (threshold: {max_distance_km} km)...")
    
    # Convert trail coordinates to LineString for efficient distance calculation
    trail_line = LineString([(coord[0], coord[1]) for coord in trail_coords])
    
    # Initialize results
    distances = []
    beyond_threshold = []
    
    valid_stamps = stamps_df.dropna(subset=['latitude', 'longitude'])
    valid_stamps = valid_stamps[(valid_stamps['latitude'] != 0) & (valid_stamps['longitude'] != 0)]
    
    print(f"Processing {len(valid_stamps)} stamps with valid coordinates...")
    
    for idx, (_, stamp) in enumerate(valid_stamps.iterrows()):
        if idx % 100 == 0:
            print(f"  Processing stamp {idx+1}/{len(valid_stamps)}...")
        
        # Create Point for stamp
        stamp_point = Point(stamp['longitude'], stamp['latitude'])
        
        # Find nearest point on trail
        nearest_point = nearest_points(stamp_point, trail_line)[1]
        
        # Calculate distance using Haversine formula
        distance_km = haversine_distance(
            stamp['latitude'], stamp['longitude'],
            nearest_point.y, nearest_point.x
        )
        
        distances.append(distance_km)
        
        # Check if beyond threshold
        if distance_km > max_distance_km:
            beyond_threshold.append({
                'place': stamp['place'],
                'town': stamp['town'],
                'latitude': stamp['latitude'],
                'longitude': stamp['longitude'],
                'distance_km': distance_km,
                'nearest_trail_lat': nearest_point.y,
                'nearest_trail_lon': nearest_point.x
            })
    
    print("✓ Distance calculation complete!")
    print(f"  - Valid stamps processed: {len(valid_stamps)}")
    print(f"  - Stamps beyond {max_distance_km}km threshold: {len(beyond_threshold)}")
    
    return beyond_threshold, distances

def create_interactive_map(trail_coords, stamps_df, wrongly_geocoded, max_distance_km):
    """Create an interactive folium map showing trail and stamps"""
    print("Creating interactive map...")
    
    # Calculate center point from trail coordinates
    trail_lats = [coord[1] for coord in trail_coords]
    trail_lons = [coord[0] for coord in trail_coords]
    center_lat = (min(trail_lats) + max(trail_lats)) / 2
    center_lon = (min(trail_lons) + max(trail_lons)) / 2
    
    # Create map
    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=8,
        tiles='OpenStreetMap'
    )
    
    # Add trail
    folium.PolyLine(
        locations=[[coord[1], coord[0]] for coord in trail_coords],
        color='blue',
        weight=4,
        opacity=0.8,
        popup='Camino Frances Trail'
    ).add_to(m)
    
    # Add stamps
    for _, stamp in stamps_df.iterrows():
        if pd.isna(stamp['latitude']) or pd.isna(stamp['longitude']):
            continue
            
        # Determine color based on whether it's wrongly geocoded
        is_wrongly_geocoded = any(
            wg['place'] == stamp['place'] and wg['town'] == stamp['town']
            for wg in wrongly_geocoded
        )
        
        color = 'red' if is_wrongly_geocoded else 'green'
        popup_text = f"""
        <b>{stamp['place']}</b><br>
        Town: {stamp['town']}<br>
        Coordinates: {stamp['latitude']:.6f}, {stamp['longitude']:.6f}<br>
        Status: {'⚠️ WRONGLY GEOCODED' if is_wrongly_geocoded else '✅ Correctly positioned'}
        """
        
        folium.CircleMarker(
            location=[stamp['latitude'], stamp['longitude']],
            radius=4,
            color=color,
            fill=True,
            fillColor=color,
            fillOpacity=0.7,
            popup=folium.Popup(popup_text, max_width=300)
        ).add_to(m)
    
    # Save map
    map_filename = 'wrongly_geocoded_stamps_map.html'
    m.save(map_filename)
    print(f"✓ Interactive map saved as '{map_filename}'")

def export_to_excel(wrongly_geocoded, max_distance_km):
    """Export wrongly geocoded stamps to Excel with formatting"""
    if not wrongly_geocoded:
        print("No wrongly geocoded stamps to export")
        return
    
    print("Creating Excel export...")
    
    # Create DataFrame
    df = pd.DataFrame(wrongly_geocoded)
    
    # Add google_search_name column
    df['google_search_name'] = df['place'] + ', ' + df['town'] + ', Spain'
    
    # Sort by distance (most distant first)
    df = df.sort_values('distance_km', ascending=False)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Stamps Beyond {max_distance_km}km"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Pilgrim Stamps Beyond {max_distance_km}km from Camino Frances Trail"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add summary
    ws.merge_cells('A2:H2')
    summary_cell = ws['A2']
    summary_cell.value = f"Total stamps found: {len(wrongly_geocoded)} | Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_cell.font = Font(italic=True, size=10)
    summary_cell.alignment = Alignment(horizontal="center")
    
    # Add headers
    headers = [
        'Place Name',
        'Town',
        'Current Latitude',
        'Current Longitude', 
        'Distance from Trail (km)',
        'Nearest Trail Latitude',
        'Nearest Trail Longitude',
        'Google Search Name'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    for row_idx, (_, stamp) in enumerate(df.iterrows(), 5):
        ws.cell(row=row_idx, column=1, value=stamp['place']).border = thin_border
        ws.cell(row=row_idx, column=2, value=stamp['town']).border = thin_border
        ws.cell(row=row_idx, column=3, value=round(stamp['latitude'], 6)).border = thin_border
        ws.cell(row=row_idx, column=4, value=round(stamp['longitude'], 6)).border = thin_border
        ws.cell(row=row_idx, column=5, value=round(stamp['distance_km'], 2)).border = thin_border
        ws.cell(row=row_idx, column=6, value=round(stamp['nearest_trail_lat'], 6)).border = thin_border
        ws.cell(row=row_idx, column=7, value=round(stamp['nearest_trail_lon'], 6)).border = thin_border
        ws.cell(row=row_idx, column=8, value=stamp['google_search_name']).border = thin_border
        
        # Apply formatting to all cells in the row
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
    
    # Auto-adjust column widths
    for col in range(1, 9):
        max_length = 0
        column_letter = ws.cell(row=4, column=col).column_letter
        
        for row in range(4, len(wrongly_geocoded) + 5):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save Excel file
    excel_filename = f'wrongly_geocoded_stamps_{max_distance_km}km.xlsx'
    wb.save(excel_filename)
    print(f"✓ Excel file saved as '{excel_filename}'")
    
    return excel_filename

def main():
    """Main function to run the analysis"""
    print("Stamp Distance Analysis Tool")
    print("=" * 40)
    
    # Parse arguments
    args = parse_arguments()
    max_distance_km = args.distance
    input_file = args.input_file
    print(f"Distance threshold: {max_distance_km} km\n")
    
    # Load data
    print("Loading data files...")
    trail_coords = load_trail_data('camino_frances_main_trail_simple.geojson')
    stamps_df = load_stamps_data(input_file)
    print("Data loaded successfully!")
    print(f"  - Trail coordinates: {len(trail_coords)} points")
    print(f"  - Stamps: {len(stamps_df)} records\n")
    
    # Calculate distances
    wrongly_geocoded, all_distances = calculate_distances_to_trail(
        stamps_df, trail_coords, max_distance_km
    )
    
    if not wrongly_geocoded:
        print("✅ All stamps are within the distance threshold!")
        return
    
    # Display results
    print("\nWrongly geocoded stamps found:")
    print(f"  - Min distance: {min(wg['distance_km'] for wg in wrongly_geocoded):.2f} km")
    print(f"  - Max distance: {max(wg['distance_km'] for wg in wrongly_geocoded):.2f} km")
    print(f"  - Mean distance: {np.mean([wg['distance_km'] for wg in wrongly_geocoded]):.2f} km")
    
    print("\nMost distant stamps:")
    sorted_stamps = sorted(wrongly_geocoded, key=lambda x: x['distance_km'], reverse=True)
    for stamp in sorted_stamps[:5]:
        print(f"  - {stamp['place']} in {stamp['town']}: {stamp['distance_km']:.2f} km")
    
    # Create map
    print("\nCreating interactive map...")
    create_interactive_map(trail_coords, stamps_df, wrongly_geocoded, max_distance_km)
    print("\nMap creation complete - Excel export pending...")
    
    # Export to Excel
    excel_file = export_to_excel(wrongly_geocoded, max_distance_km)
    print(f"\n✅ Analysis complete!")
    print(f"📊 Results exported to: {excel_file}")
    print(f"🗺️  Interactive map: wrongly_geocoded_stamps_map.html")

if __name__ == "__main__":
    main()
